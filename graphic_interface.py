from tkinter import *
from tkinter.ttk import Progressbar
import openpyxl
import requests
from bs4 import BeautifulSoup
import time


class Sc:
    def __init__(self):
        self.read_settings = openpyxl.open('settings.xlsx', read_only=True)
        self.sheet_settings = self.read_settings.active
        self.all_tags = []
        self.read_shop_url = openpyxl.open('shop_url.xlsx', read_only=True)
        self.sheet_shop_url = self.read_shop_url.active
        self.new_book = openpyxl.Workbook()
        self.new_sheet = self.new_book.active

    def get_html_tags(self):
        for row in self.sheet_settings.iter_rows(min_row=self.all_row_col[0], max_row=self.all_row_col[1], min_col=2,
                                                 max_col=4):
            for cell in row:
                read_cell = cell.value
                self.all_tags.append(read_cell)
        return self.all_tags

    def get_result(self):
        tag1 = 0
        tag2 = 1
        tag3 = 2
        jump = self.all_row_col[3] + 1
        headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}
        for row in self.sheet_shop_url.iter_rows(min_row=self.all_row_col[0], max_row=self.all_row_col[1],
                                                 min_col=self.all_row_col[2], max_col=self.all_row_col[3]):
            for cell in row:
                read_cell = cell.value
                if read_cell == 'x':
                    self.new_sheet.cell(row=self.all_row_col[0], column=self.all_row_col[2]).value = 'x'
                else:
                    url = read_cell
                    print(url)
                    self.label_error.configure(text=url)
                    self.label_error.update()
                    try:
                        source = requests.get(url, headers=headers)
                        main_text = source.content.decode()
                        soup = BeautifulSoup(main_text, features="html.parser")
                        price = soup.find(self.all_tags[tag1], {self.all_tags[tag2]: self.all_tags[tag3]})
                        price = price.text

                        price = price.replace(' ', '')
                        price = price.replace('грн', '')
                        price = price.replace('грн.', '')
                        price = price.replace('\xa0', '')
                        price = price.replace('.', '')
                        price = price.replace('₴', '')
                        price = price.replace('\n', '')
                        price = price.replace('\t', '')
                        price = price.replace('.', '')
                        price = price.replace(':', '')
                        price = price.replace('Цена', '')

                        formula = f'=HYPERLINK("{url}";"{price}")'
                        self.new_sheet.cell(row=self.all_row_col[0], column=self.all_row_col[2]).value = formula
                    except BaseException:
                        print('Error: ' + url)
                        url_not_work = f'=HYPERLINK("{url}";"!404!")'
                        self.new_sheet.cell(row=self.all_row_col[0], column=self.all_row_col[2]).value = url_not_work

                self.all_row_col[2] += 1

                if self.all_row_col[2] == jump:
                    self.all_row_col[0] += 1
                    self.all_row_col[2] = 2
                    tag1 += 3
                    tag2 += 3
                    tag3 += 3

    def save(self):
        self.new_book.save('class_price_url.xlsx')
        self.new_book.close()


class Window(Sc):
    def __init__(self):
        super().__init__()
        self.canvas = Tk()
        self.canvas.title('Перевірка цін з конкурентами')
        self.canvas.geometry('600x300+500+200')
        self.canvas.resizable(False, False)
        self.label_s_r = Label(self.canvas, text='Start row')
        default_start_row = StringVar(self.canvas, value='4')
        self.entry_s_r = Entry(self.canvas, textvariable=default_start_row)
        self.label_e_r = Label(self.canvas, text='End row')
        default_end_row = StringVar(self.canvas, value='26')
        self.entry_e_r = Entry(self.canvas, textvariable=default_end_row)
        self.label_error = Label(self.canvas)
        self.label_s_c = Label(self.canvas, text='Start col')
        default_start_col = StringVar(self.canvas, value='2')
        self.entry_s_c = Entry(self.canvas, textvariable=default_start_col)
        self.label_e_c = Label(self.canvas, text='End col')
        default_end_col = StringVar(self.canvas, value='16')
        self.entry_e_c = Entry(self.canvas, textvariable=default_end_col)

        self.all_row_col = []

    def run(self):
        self.draw_widgets()
        self.canvas.mainloop()

    def draw_widgets(self):
        self.label_s_r.pack()
        self.entry_s_r.pack()
        self.label_e_r.pack()
        self.entry_e_r.pack()
        self.label_s_c.pack()
        self.entry_s_c.pack()
        self.label_e_c.pack()
        self.entry_e_c.pack()

        Button(self.canvas, text='Go!', command=self.button_action).pack(pady=10)
        self.label_error.pack()

    def button_action(self):
        self.all_tags.clear()
        self.all_row_col.clear()
        self.label_error.config(text='')
        self.get_row_and_col()
        if self.all_row_col:
            self.get_html_tags()
            self.get_result()
        self.save()

    def get_row_and_col(self):
        try:
            self.all_row_col.append(abs(int(self.entry_s_r.get())))
            self.all_row_col.append(int(self.entry_e_r.get()))
            self.all_row_col.append(int(self.entry_s_c.get()))
            self.all_row_col.append(int(self.entry_e_c.get()))
            if self.all_row_col[0] > self.all_row_col[1] and self.all_row_col[2] > self.all_row_col[3]: # прочекати
                self.label_error.config(text='Невірно задані значення')
                self.all_row_col.clear()
        except ValueError:
            self.label_error.config(text='Введи число')
            self.all_row_col.clear()

        print(self.all_row_col)


if __name__ == "__main__":
    window = Window()
    window.run()
