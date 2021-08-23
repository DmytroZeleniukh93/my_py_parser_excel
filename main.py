import openpyxl
import requests
from bs4 import BeautifulSoup


class ScrapPrices:
    replace_list = (' ', 'грн', '.', 'грн.', '\xa0', '₴', '\n', '\t', 'Цена', ':')
    all_tags = []

    def __init__(self):
        self.read_settings = openpyxl.open('settings.xlsx', read_only=True)
        self.sheet_settings = self.read_settings.active
        self.read_shop_url = openpyxl.open('shop_url.xlsx', read_only=True)
        self.shop_url = self.read_shop_url.active
        self.new_book = openpyxl.Workbook()
        self.new_sheet = self.new_book.active

    def get_tags(self, start_row, end_row):
        for row in self.sheet_settings.iter_rows(min_row=start_row, max_row=end_row, min_col=2, max_col=4):
            for cell in row:
                read_cell = cell.value
                self.all_tags.append(read_cell)
                # print(read_cell)  # добавити фічу забрати пробіл в кінці ексель файлу

    def get_result(self, start_row, end_row, start_col, end_col):
        tag1 = 0
        tag2 = 1
        tag3 = 2
        v = self.res
        vv = v
        headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}
        for row in self.shop_url.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            for cell in row:
                read_cell = cell.value
                self.pb.configure(value=v)
                self.pb.update()
                if read_cell == 'x':
                    self.new_sheet.cell(row=start_row, column=start_col).value = 'x'
                else:
                    url = read_cell
                   # print(url)
                    self.label_url.configure(text=url)
                    self.label_url.update()

                    try:
                        source = requests.get(url, headers=headers)
                        main_text = source.content.decode()
                        soup = BeautifulSoup(main_text, features="html.parser")
                        price = soup.find(self.all_tags[tag1], {self.all_tags[tag2]: self.all_tags[tag3]})
                        price = self.__get_clean_price(price.text)
                        to_write = f'=HYPERLINK("{url}";"{price}")'
                        self.new_sheet.cell(row=start_row, column=start_col).value = to_write
                    except Exception:
                        #print('Error: ' + url)
                        to_write = f'=HYPERLINK("{url}";"!404!")'
                        self.new_sheet.cell(row=start_row, column=start_col).value = to_write

                start_col += 1
                v += vv
                if start_col == end_col + 1:
                    start_row += 1
                    start_col = 2
                    tag1 += 3
                    tag2 += 3
                    tag3 += 3

    def __get_clean_price(self, price):
        for to_replace in self.replace_list:
            price = price.replace(to_replace, '')
        return price

    def __close_files(self):
        self.read_settings.close()
        self.read_shop_url.close()
        self.new_book.close()

    def save(self):
        self.new_book.save('new_price.xlsx')
        self.new_book.close()
        self.__close_files()
